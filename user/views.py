from django.shortcuts import render,redirect
from django.contrib import messages
from django.contrib.auth import authenticate,login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from .forms import UserRegisterForm
from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.template.loader import get_template
from django.template import Context
import lstm_emotion
import json
import graph
import rating
j=0

def gr(name,j):
    user1=val()
    with open('val.txt') as f:
            lines = f.readlines()
            print(lines)
        
    lines[j-1]=lines[j-1][:-1]+" "+str(user1)+"\n"
    print(lines)
    
    # with open('val.txt', 'w') as f:
    #     for line in lines:
    #         f.write(line)
    with open('review.txt') as f:
        lines = f.readlines()
    print((lines))
    print((type(lines[j-1])))
    line=lines[j-1].split("{")
    print(line)
    rl=len(line)
    for l in line:
        print(l)
    return line
def validate(name,j):
    # import module
    error=0
    user1=val()
    with open('val.txt') as f:
        lines = f.readlines()
        print(lines)
    print("lines to compare",lines[j-1])
    if user1 in lines[j-1]:
        print("not allowed to review")
        error=1
        line=[]
    else:
        with open('val.txt') as f:
            lines = f.readlines()
            print(lines)
        
        lines[j-1]=lines[j-1][:-1]+" "+str(user1)+"\n"
        print(lines)
        
        # with open('val.txt', 'w') as f:
        #     for line in lines:
        #         f.write(line)
        with open('review.txt') as f:
            lines = f.readlines()
        print((lines))
        print((type(lines[j-1])))
        line=lines[j-1].split("{")
        print(line)
        rl=len(line)
        for l in line:
            print(l)
    return error,line

def loadrating():
    global j
    with open("joy.json", 'r') as f:
        score1 = json.load(f)
        
    joy = score1
    with open("fear.json", 'r') as f:
        score2 = json.load(f)
    fear = score2
    with open("anger.json", 'r') as f:
        score3 = json.load(f)
        
    anger = score3
    with open("sadness.json", 'r') as f:
        score4 = json.load(f)
    sadness = score4
    with open("neutral.json", 'r') as f:
        score5 = json.load(f)
    neutral = score5
    
    gr={"joy":joy[j-1],"fear":fear[j-1],"anger":anger[j-1],"sadness":sadness[j-1],"neutral":neutral[j-1]}
    
    graph.main(gr)
    rat=[]
    i = 0
    print("graph values",gr ,"and j is",j)
    
    for s in range(8):
        a=joy[i]
        b=fear[i]
        c=anger[i]
        d=sadness[i]
        e=neutral[i]
        #if a == b == c == d == e:
         #   rat.append("4")
        if a > b and a > c and a > d and a > e:
            rat.append("5")
        elif b>a and b>c and b>d and b>e:
            rat.append("1")
        elif c>a and c>b and c>d and c>e:
            rat.append("2")
        elif d>a and d>c and d>b and d>e:
            rat.append("3")
        elif e>a and e>c and e>d and e>b:
            rat.append("4")
        else:
            rat.append("4")
        i=i+1
    
    print("rat is",rat)
    return rat
def rvcount():
    with open('review.txt') as f:
            lines = f.readlines()
    print((lines))
    print(type(lines))
    lens=[]
    for line in lines:
        line=line.split("{")
        print(line)
        rl=len(line)
        lens.append(rl)
    print("review counts:",lens)
    return lens
    

##################################################################
####################index#######################################
def index(request):
    cnt=rvcount()
    print("cnt in index is",cnt)
    rat=loadrating()      
    return render(request, 'user/index.html',{'title':'index',"cnt1":cnt[0] ,"cnt2":cnt[1],"cnt3":cnt[2],"cnt4":cnt[3],"cnt5":cnt[4],"cnt6":cnt[5],"cnt7":cnt[6],"cnt8":cnt[7],"pr1":rat[0] ,"pr2":rat[1],"pr3":rat[2],"pr4":rat[3],"pr5":rat[4],"pr6":rat[5],"pr7":rat[6],"pr8":rat[7] })

########################################################################
########### register here #####################################
def movie(request):
    return render(request, 'user/movie.html',{'title':'movies'})


def sub(request):
# =============================================================================
     global j
     user1=val()
     with open('val.txt') as f:
         lines = f.readlines()
         print(lines)
     print("lines to compare",lines[j-1])
     if user1 in lines[j-1]:
         print("not allowed to review")
         text="you have recored your review once!! Please visit homepage for more movie"
     else:
# =============================================================================
            comm = request.GET["subject"]
            res=lstm_emotion.main(comm)
            if res=='joy':
                text="System Detected Your Review emotion To Be Joy"
                rating.main(0,j)
            elif res=='anger':
                text="System Detected Your Review emotion To Be Anger"
                rating.main(1,j)
            elif res=='neutral':
                text="System Detected Your Review emotion To Be Neutral"
                rating.main(2,j)
            elif res=='fear':
                text="System Detected Your Review emotion To Be Fear"
                rating.main(3,j)
            elif res=='sadness':
                text="System Detected Your Review emotion To Be Sad"
                rating.main(4,j)
            print("comment is",comm)
            loadrating()
            print("j is",j)
            with open('review.txt') as f:
                    lines = f.readlines()
                    print(lines)
                
            lines[j-1]=lines[j-1][:-1]+"{"+str(comm)+"\n"
            print(lines)
            with open('review.txt', 'w') as f:
                for line in lines:
                    f.write(line)
                    
            with open('val.txt') as f:
                lines = f.readlines()
                print(lines)
            
            lines[j-1]=lines[j-1][:-1]+" "+str(user1)+"\n"
            print(lines)
            
            with open('val.txt', 'w') as f:
                for line in lines:
                    f.write(line)

     cnt=rvcount()
     return render(request, 'user/rating.html',{'title':'rating',"cnt":cnt,"text":text})

def val():
    # import module
    
    import openpyxl
      
    # load excel with its path
    wrkbk = openpyxl.load_workbook("login.xlsx")
      
    sh = wrkbk.active
    cell_obj = sh.cell(row=1, column=1)
    print("user is", cell_obj.value)
    return cell_obj.value
    


    
    

def xmen(request):
    global j
    print("movie choosen is xmen")
    name="xmen"
    j=1
    l=gr(name,j)
    loadrating()
    # if err==1:
    #     messages.info(request, f'you have reviwed this movie already')
    #     return redirect('index')
    # else:
    return render(request, 'user/rating.html',{'title':name,"l":l,"name":name})
def spm2(request):
    global j
    print("movie choosen is spm2")
    name="spm2"
    j=2
    l=gr(name,j)
    loadrating()
    # if err==1:
    #     messages.info(request, f'you have reviwed this movie already')
    #     return redirect('index')
    # else:
    return render(request, 'user/rating.html',{'title':name,"l":l,"name":name})
def spm3(request):
    global j
    print("movie choosen is spm3")
    name="spm3"
    j=3
    l=gr(name,j)
    loadrating()
    # if err==1:
    #     messages.info(request, f'you have reviwed this movie already')
    #     return redirect('index')
    # else:
    return render(request, 'user/rating.html',{'title':name,"l":l,"name":name})
def valkyrie(request):
    global j
    print("movie choosen is valkyrie")
    name="valkyrie"
    j=4
    l=gr(name,j)
    loadrating()
    # if err==1:
    #     messages.info(request, f'you have reviwed this movie already')
    #     return redirect('index')
    # else:
    return render(request, 'user/rating.html',{'title':name,"l":l,"name":name})
def gladiator(request):
    global j
    print("movie choosen is gladiator")
    name="gladiator"
    j=5
    l=gr(name,j)
    loadrating()
    # if err==1:
        
    #     messages.info(request, f'you have reviwed this movie already')
    #     return redirect('index')
    # else:
    return render(request, 'user/rating.html',{'title':name,"l":l,"name":name})
def iceage(request):
    global j
    print("movie choosen is iceage")
    name="iceage"
    j=6
    l=gr(name,j)
    loadrating()
    # if err==1:
    #     messages.info(request, f'you have reviwed this movie already')
    #     return redirect('index')
    # else:
    return render(request, 'user/rating.html',{'title':name,"l":l,"name":name})
def transformers(request):
    global j
    print("movie choosen is transformers")
    name="transformers"
    j=7
    l=gr(name,j)
    loadrating()
    # if err==1:
    #     messages.info(request, f'you have reviwed this movie already')
    #     return redirect('index')
    # else:
    return render(request, 'user/rating.html',{'title':name,"l":l,"name":name})
def magneto(request):
    global j
    print("movie choosen is magneto")
    name="magneto"
    j=8
    l=gr(name,j)
    loadrating()
    # if err==1:
    #     messages.info(request, f'you have reviwed this movie already')
    #     return redirect('index')
    # else:
    print("reviews are",l)
    return render(request, 'user/rating.html',{'title':name,"l":l,"name":name})
    
def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST) or None
        if form.is_valid():
           
            username = request.POST.get('username')
            #########################mail####################################
            htmly = get_template('user/Email.html')
            d = { 'username': username }
            subject, from_email, to = 'hello', 'from@example.com', 'to@emaple.com'
            html_content = htmly.render(d)
            msg = EmailMultiAlternatives(subject, html_content, from_email, [to])
            msg.attach_alternative(html_content, "text/html")
            try:
                msg.send()
            except:
                print("error in sending mail")
            ##################################################################
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Your account has been created! You are now able to log in')
            return redirect('login')
    else:
        form = UserRegisterForm()
    return render(request, 'user/register.html', {'form': form,'title':'reqister here'})

###################################################################################
################login forms###################################################


    
def Login(request):
    if request.method == 'POST':

        #AuthenticationForm_can_also_be_used__

        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            form = login(request,user)
            messages.success(request, f' wecome {username} !!')
            # import openpyxl module
            import openpyxl
             
            wb = openpyxl.Workbook()
              
            sheet = wb.active
              
            c1 = sheet.cell(row = 1, column = 1)
             
            c1.value = username
            wb.save("login.xlsx")
            cnt=rvcount()
            rat=loadrating()
            return render(request, 'user/index.html', {'form':form,'title':'homepage',"cnt1":cnt[0] ,"cnt2":cnt[1],"cnt3":cnt[2],"cnt4":cnt[3],"cnt5":cnt[4],"cnt6":cnt[5],"cnt7":cnt[6],"cnt8":cnt[7],"pr1":rat[0] ,"pr2":rat[1],"pr3":rat[2],"pr4":rat[3],"pr5":rat[4],"pr6":rat[5],"pr7":rat[6],"pr8":rat[7]})
            #return redirect('index',{"cnt":cnt})
        else:
            messages.info(request, f'account does not exit plz sign in')
    form = AuthenticationForm()
    return render(request, 'user/login.html', {'form':form,'title':'log in'})
